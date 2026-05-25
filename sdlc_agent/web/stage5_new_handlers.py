"""New Stage 5 handlers for 4-button workflow with LLM-powered test generation"""
from pathlib import Path
import subprocess
import sys
import json
from datetime import datetime, timezone

from ..models import StoryBacklog, PullRequest, UserStory
from ..testing_assets import write_manual_tests_xlsx, write_playwright_suite
from ..integrations.anthropic_client import MockClaudeClient


def _generate_detailed_test_cases_with_llm(story: UserStory, llm: MockClaudeClient) -> list[dict]:
    """Use LLM to generate detailed test cases with steps for a story."""
    system_prompt = """You are a QA engineer creating detailed manual test cases.
For each acceptance criterion, generate a comprehensive test case with:
- Test steps (numbered, clear actions)
- Test data (specific inputs to use)
- Expected results (what should happen)
- Priority (High/Medium/Low)
- Type (Functional/UI/API/Performance/Security)

Return JSON array with this structure:
[
  {
    "tc_id": "TC-001",
    "title": "Test case title",
    "steps": ["1. Action", "2. Action", "3. Action"],
    "test_data": "Specific data to use",
    "expected_result": "What should happen",
    "priority": "High",
    "type": "Functional"
  }
]"""

    user_prompt = f"""Story: {story.id}
As a: {story.persona}
I want: {story.want}
So that: {story.so_that}

Acceptance Criteria:
{chr(10).join(f"- {ac}" for ac in story.acceptance_criteria)}

Generate detailed test cases for each acceptance criterion."""

    result = llm.complete_json(
        system=system_prompt,
        user=user_prompt,
        max_tokens=4096,
        temperature=0.3
    )

    # Debug logging
    import sys
    from pathlib import Path
    debug_file = Path(__file__).resolve().parents[2] / "stage5_debug.log"
    with open(debug_file, "a") as f:
        f.write(f"  LLM raw result for {story.id}: {result}\n")
        f.write(f"  Result type: {type(result)}, Is list: {isinstance(result, list)}\n")

    if result and isinstance(result, list):
        return result
    return []


def _generate_playwright_script_with_llm(story: UserStory, llm: MockClaudeClient) -> str:
    """Use LLM to generate intelligent Playwright test script."""
    system_prompt = """You are a test automation engineer writing Playwright TypeScript tests.
Generate a complete, runnable Playwright test file that:
- Uses proper selectors (data-testid, role-based, or CSS)
- Includes appropriate waits (waitForSelector, waitForLoadState)
- Has clear assertions
- Handles edge cases
- Follows Playwright best practices

Return ONLY the TypeScript code, no markdown formatting."""

    user_prompt = f"""Story: {story.id}
As a: {story.persona}
I want: {story.want}
So that: {story.so_that}

Acceptance Criteria:
{chr(10).join(f"- {ac}" for ac in story.acceptance_criteria)}

Generate a Playwright test file that covers all acceptance criteria."""

    result = llm.complete_json(
        system=system_prompt,
        user=user_prompt,
        max_tokens=4096,
        temperature=0.3
    )

    if isinstance(result, dict) and "code" in result:
        return result["code"]
    elif isinstance(result, str):
        return result
    return ""


def generate_manual_tests(run_id: str, root: Path, manual_tests_dir: Path) -> dict:
    """Step 1: Generate manual test cases Excel with LLM-enhanced details"""
    try:
        runs_dir = root / "runs"
        rd = runs_dir / run_id

        # Read backlog
        backlog_path = rd / "02_backlog.json"
        if not backlog_path.exists():
            return {"error": "Backlog not found. Run Stage 2 first."}, 404

        backlog = StoryBacklog.model_validate_json(backlog_path.read_text(encoding="utf-8"))

        # Initialize LLM client
        llm = MockClaudeClient()

        # Debug logging to file (since stderr isn't showing up)
        debug_file = root / "stage5_debug.log"
        with open(debug_file, "a") as f:
            import datetime
            f.write(f"\n[{datetime.datetime.now()}] Stage5 Manual Tests Called\n")
            f.write(f"  LLM Backend: {llm.backend}\n")
            f.write(f"  LLM Is Live: {llm.is_live}\n")
            f.write(f"  Run ID: {run_id}\n")

        # Generate Excel with LLM-enhanced test cases if LLM is available
        manual_tests_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = manual_tests_dir / f"{run_id}_manual_tests.xlsx"

        if llm.is_live:
            # LLM-powered: Generate detailed test cases for each story
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            summary = wb.active
            summary.title = "Summary"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="305496")

            summary.append(["TC ID", "Story", "Title", "Priority", "Type", "LLM Generated"])
            for col in range(1, 7):
                cell = summary.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

            detail = wb.create_sheet("Test Cases")
            detail.append(["TC ID", "Story", "Title", "Steps", "Test Data", "Expected Result", "Priority", "Type"])
            for col in range(1, 9):
                cell = detail.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

            row = 2
            for story in backlog.stories:
                test_cases = _generate_detailed_test_cases_with_llm(story, llm)

                # Debug: Log what we got from LLM
                with open(debug_file, "a") as f:
                    f.write(f"  Story {story.id}: Got {len(test_cases)} test cases\n")
                    f.write(f"  Raw result type: {type(test_cases)}\n")
                    if test_cases:
                        f.write(f"  First test case: {test_cases[0]}\n")

                for tc in test_cases:
                    summary.append([
                        tc.get("tc_id", f"TC-{row:03d}"),
                        story.id,
                        tc.get("title", ""),
                        tc.get("priority", "Medium"),
                        tc.get("type", "Functional"),
                        "✓"
                    ])

                    steps_text = "\n".join(tc.get("steps", []))
                    detail.append([
                        tc.get("tc_id", f"TC-{row:03d}"),
                        story.id,
                        tc.get("title", ""),
                        steps_text,
                        tc.get("test_data", ""),
                        tc.get("expected_result", ""),
                        tc.get("priority", "Medium"),
                        tc.get("type", "Functional")
                    ])
                    row += 1

            wb.save(xlsx_path)
            test_case_count = row - 2
            backend_info = f" (LLM: {llm.backend})"
        else:
            # Fallback: Template-based generation
            write_manual_tests_xlsx(backlog, xlsx_path)
            test_case_count = sum(len(s.acceptance_criteria) for s in backlog.stories)
            backend_info = " (Template-based)"

        story_count = len(backlog.stories)

        return {
            "success": True,
            "run_id": run_id,
            "file_path": str(xlsx_path.relative_to(root)),
            "story_count": story_count,
            "test_case_count": test_case_count,
            "message": f"Generated {test_case_count} detailed test cases for {story_count} stories{backend_info}"
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": str(e)}, 500


def generate_automation_scripts(run_id: str, root: Path, automation_dir: Path) -> dict:
    """Step 2: Generate Playwright automation scripts with LLM-powered intelligence"""
    try:
        runs_dir = root / "runs"
        rd = runs_dir / run_id

        # Read backlog
        backlog_path = rd / "02_backlog.json"
        if not backlog_path.exists():
            return {"error": "Backlog not found."}, 404

        backlog = StoryBacklog.model_validate_json(backlog_path.read_text(encoding="utf-8"))

        # Initialize LLM client
        llm = MockClaudeClient()

        # Generate Playwright suite
        automation_dir.mkdir(parents=True, exist_ok=True)
        playwright_dir = automation_dir / run_id
        playwright_dir.mkdir(parents=True, exist_ok=True)
        tests_dir = playwright_dir / "tests"
        tests_dir.mkdir(parents=True, exist_ok=True)

        if llm.is_live:
            # LLM-powered: Generate intelligent Playwright scripts
            spec_paths = []
            for story in backlog.stories:
                script_content = _generate_playwright_script_with_llm(story, llm)
                if script_content:
                    # Clean story.want to create valid filename
                    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in story.want)
                    safe_name = safe_name.replace(" ", "_").lower()[:50]
                    spec_file = tests_dir / f"{safe_name}.spec.ts"
                    spec_file.write_text(script_content, encoding="utf-8")
                    spec_paths.append(spec_file)

            # Write configuration files
            (playwright_dir / "package.json").write_text('{"devDependencies": {"@playwright/test": "^1.40.0"}}', encoding="utf-8")
            (playwright_dir / "playwright.config.ts").write_text(
                "export default { testDir: './tests', timeout: 30000 };",
                encoding="utf-8"
            )

            backend_info = f" (LLM: {llm.backend})"
            script_count = len(spec_paths)
            scripts = [str(p.relative_to(root)) for p in spec_paths]
        else:
            # Fallback: Template-based generation
            pw_info = write_playwright_suite(backlog, playwright_dir)
            backend_info = " (Template-based)"
            script_count = len(pw_info["specs"])
            scripts = [str(p.relative_to(root)) for p in pw_info["specs"]]

        return {
            "success": True,
            "run_id": run_id,
            "playwright_dir": str(playwright_dir.relative_to(root)),
            "script_count": script_count,
            "scripts": scripts,
            "message": f"Generated {script_count} intelligent Playwright test scripts{backend_info}"
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": str(e)}, 500


def execute_tests(run_id: str, root: Path, automation_dir: Path, results_dir: Path) -> dict:
    """Step 3: Execute Playwright tests"""
    try:
        playwright_dir = automation_dir / run_id
        if not playwright_dir.exists():
            return {"error": "Automation scripts not found. Generate them first."}, 404

        results_dir.mkdir(parents=True, exist_ok=True)

        # Run Playwright tests
        import shutil
        npx = shutil.which("npx")
        if not npx:
            return {
                "error": "npx not found. Install Node.js and Playwright.",
                "executed": False
            }, 400

        # Run with JSON reporter
        proc = subprocess.run(
            [npx, "playwright", "test", "--reporter=json"],
            cwd=str(playwright_dir),
            capture_output=True,
            text=True,
            timeout=180,
        )

        # Save results
        result_file = results_dir / f"{run_id}_results.json"
        html_file = results_dir / f"{run_id}_results.html"
        log_file = results_dir / f"{run_id}_execution.log"

        log_file.write_text(
            f"STDOUT:\n{proc.stdout}\n\nSTDERR:\n{proc.stderr}",
            encoding="utf-8"
        )

        # Parse results
        try:
            if proc.stdout:
                results = json.loads(proc.stdout)
                result_file.write_text(json.dumps(results, indent=2), encoding="utf-8")

                # Count pass/fail
                stats = results.get("stats", {})
                passed = stats.get("passed", 0)
                failed = stats.get("failed", 0)
                total = passed + failed
            else:
                passed, failed, total = 0, 0, 0
                results = {"error": "No output from Playwright"}
        except json.JSONDecodeError:
            # Fallback: parse text output
            passed = proc.stdout.count("passed")
            failed = proc.stdout.count("failed")
            total = passed + failed
            results = {"stdout": proc.stdout, "stderr": proc.stderr}

        return {
            "success": True,
            "executed": True,
            "run_id": run_id,
            "exit_code": proc.returncode,
            "passed": passed,
            "failed": failed,
            "total": total,
            "has_failures": failed > 0,
            "result_file": str(result_file.relative_to(root)),
            "log_file": str(log_file.relative_to(root)),
            "message": f"Executed {total} tests: {passed} passed, {failed} failed"
        }
    except subprocess.TimeoutExpired:
        return {"error": "Test execution timed out (3 minutes)"}, 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": str(e)}, 500


def heal_tests(run_id: str, root: Path, automation_dir: Path, results_dir: Path) -> dict:
    """Step 4: Analyze failures and heal tests using AI"""
    try:
        # Read test results
        result_file = results_dir / f"{run_id}_results.json"
        if not result_file.exists():
            return {"error": "Test results not found. Execute tests first."}, 404

        results = json.loads(result_file.read_text(encoding="utf-8"))

        # Extract failures
        failures = []
        suites = results.get("suites", [])
        for suite in suites:
            for spec in suite.get("specs", []):
                for test in spec.get("tests", []):
                    for result in test.get("results", []):
                        if result.get("status") == "failed":
                            failures.append({
                                "test": test.get("title"),
                                "file": suite.get("file"),
                                "error": result.get("error", {}).get("message", "Unknown error")
                            })

        if not failures:
            return {
                "success": True,
                "message": "No failures to heal - all tests passed!",
                "healed_count": 0
            }

        # Initialize LLM client for intelligent healing
        llm = MockClaudeClient()

        # Analyze failures and suggest fixes using LLM
        healing_suggestions = []
        for failure in failures:
            suggestion = analyze_failure(failure, llm)
            healing_suggestions.append(suggestion)

        llm_powered = llm.is_live
        backend_info = f" (LLM: {llm.backend})" if llm_powered else " (Pattern-based)"

        # Save healing report
        healing_report = results_dir / f"{run_id}_healing_report.json"
        healing_report.write_text(
            json.dumps({
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "failures_analyzed": len(failures),
                "suggestions": healing_suggestions
            }, indent=2),
            encoding="utf-8"
        )

        return {
            "success": True,
            "run_id": run_id,
            "failures_found": len(failures),
            "suggestions": healing_suggestions[:5],  # Top 5
            "healing_report": str(healing_report.relative_to(root)),
            "llm_powered": llm_powered,
            "message": f"Analyzed {len(failures)} failures and generated intelligent healing suggestions{backend_info}"
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": str(e)}, 500


def analyze_failure(failure: dict, llm: MockClaudeClient = None) -> dict:
    """Analyze a single test failure using LLM for intelligent diagnosis"""
    test_name = failure["test"]
    error_msg = failure["error"]
    file_path = failure.get("file", "")

    if llm and llm.is_live:
        # LLM-powered analysis
        system_prompt = """You are a test automation expert analyzing test failures.
Given a test name, file path, and error message, diagnose the root cause and provide:
1. The specific issue (be precise)
2. A detailed fix suggestion with code examples
3. The fix type category

Return JSON with this structure:
{
  "test": "test name",
  "issue": "specific issue identified",
  "suggestion": "detailed fix with code example",
  "fix_type": "timing|selector|assertion|logic|setup|unknown"
}"""

        user_prompt = f"""Test: {test_name}
File: {file_path}
Error: {error_msg}

Analyze this test failure and provide a specific fix."""

        result = llm.complete_json(
            system=system_prompt,
            user=user_prompt,
            max_tokens=1024,
            temperature=0.2
        )

        if isinstance(result, dict) and "issue" in result:
            result["test"] = test_name
            result["llm_powered"] = True
            return result

    # Fallback: Pattern matching for common failures
    error_lower = error_msg.lower()
    if "timeout" in error_lower or "exceeded" in error_lower:
        return {
            "test": test_name,
            "issue": "Timeout",
            "suggestion": "Increase timeout or add explicit waits: await page.waitForLoadState('networkidle')",
            "fix_type": "timing",
            "llm_powered": False
        }
    elif "locator" in error_lower or "selector" in error_lower or "not found" in error_lower:
        return {
            "test": test_name,
            "issue": "Element not found",
            "suggestion": "Update selector or add wait: await page.waitForSelector('your-selector', { state: 'visible' })",
            "fix_type": "selector",
            "llm_powered": False
        }
    elif "expected" in error_lower or "assertion" in error_lower:
        return {
            "test": test_name,
            "issue": "Assertion failed",
            "suggestion": "Verify expected value matches actual. Check data setup or API response.",
            "fix_type": "assertion",
            "llm_powered": False
        }
    else:
        return {
            "test": test_name,
            "issue": "Unknown error",
            "suggestion": "Review test logic and error message. May need manual investigation.",
            "fix_type": "unknown",
            "llm_powered": False
        }
